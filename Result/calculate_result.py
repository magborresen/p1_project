from openpyxl import load_workbook

# Create a dictionary of all the frequencies
frequency_mean = {'0,25', '0,5', '1', '2', '4', '8'}


def calc_mean(list_left_ear, list_right_ear, gender):

    mean_left = 0
    mean_right = 0

    # Calculate the mean of both ears for every frequency
    frequency_mean['0.25'] = (list_left_ear[0] + list_right_ear[0]) / 2
    frequency_mean['0.5'] = (list_left_ear[1] + list_right_ear[1]) / 2
    frequency_mean['1'] = (list_left_ear[2] + list_right_ear[2]) / 2
    frequency_mean['2'] = (list_left_ear[3] + list_right_ear[3]) / 2
    frequency_mean['4'] = (list_left_ear[4] + list_right_ear[4]) / 2
    frequency_mean['8'] = (list_left_ear[5] + list_right_ear[5]) / 2

    # Calculate the total mean of the left ear
    for num in list_left_ear:
        mean_left += num

    # Calculate the total mean of the right ear
    for num in list_right_ear:
        mean_right += num

    mean_left = mean_left / 6
    mean_right = mean_right / 6

    mean = (mean_left + mean_right) / 2

    calc_age_related(mean, gender)


# Calculate the difference in age related hearing loss

def calc_age_related(mean, gender, age):

    # Choose sheetname according to gender and convert to openpyxl workbook
    if gender == "Kvinde":
        wb = load_workbook(filename="hearing_age.xlsx")
        ws = wb['Kvinder']
    elif gender == "Mand":
        wb = load_workbook("hearing_age.xlsx")
        ws = wb['Maend']

    # Compare hearing loss for every frequency relative to the users age




# Calculate the noise induced hearing loss

def calc_noise_induced(age_related):
    pass


s



# Show the result

def show_result(mean, age_related, noise_induced):
    pass
